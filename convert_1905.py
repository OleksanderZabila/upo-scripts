import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

SOURCE = r'C:\Users\Забіла\Desktop\1905.xlsx'
TARGET_TEMPLATE = r'C:\Users\Забіла\Desktop\Черкаси 19-20.05.26.xlsx'
OUTPUT = r'C:\Users\Забіла\Desktop\Черкаси_результат.xlsx'


def timedelta_to_serial(td):
    return td.total_seconds() / 86400


df19 = pd.read_excel(SOURCE, sheet_name='19.05.2026', header=None)
df20 = pd.read_excel(SOURCE, sheet_name='20.05.2026', header=None)

# combine and deduplicate (20.05 sheet duplicates night rows already in 19.05)
df_all = pd.concat([df19, df20], ignore_index=True)
df_all = df_all.drop_duplicates(subset=[1, 2])  # unique by (time, name)

# sort: early morning (< 06:00) treated as next-day so night entries come last
SHIFT_CUTOFF = 6 * 3600  # seconds

def sort_key(row):
    secs = row[1].total_seconds() if hasattr(row[1], 'total_seconds') else 0
    return secs + 86400 if secs < SHIFT_CUTOFF else secs

df_all['_sort'] = df_all.apply(sort_key, axis=1)
df_all = df_all.sort_values('_sort').drop(columns='_sort').reset_index(drop=True)

def build_rows(df):
    rows = []
    for _, row in df.iterrows():
        call_td = row[1]
        name = row[2]
        address = row[3]
        unit = row[5]
        depart_td = row[6]
        call_serial = timedelta_to_serial(call_td) if hasattr(call_td, 'total_seconds') else None
        depart_serial = timedelta_to_serial(depart_td) if hasattr(depart_td, 'total_seconds') else None
        rows.append((call_serial, name, address, call_serial, unit, depart_serial))
    return rows

data = build_rows(df_all)

# copy headers from template
tmpl = load_workbook(TARGET_TEMPLATE)
tmpl_ws = tmpl.active
headers = [cell.value for cell in tmpl_ws[1]]
col_widths = {col: dim.width for col, dim in tmpl_ws.column_dimensions.items()}

wb = Workbook()
ws = wb.active
ws.title = '19-20.05.26'

# header row
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = Font(name='Calibri', size=11, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# data rows
for ri, row in enumerate(data, 2):
    call_serial, name, address, call_serial2, unit, depart_serial = row

    c = ws.cell(row=ri, column=1, value=call_serial)
    c.number_format = 'H:MM:SS'

    ws.cell(row=ri, column=2, value=name)
    ws.cell(row=ri, column=3, value=address)

    c = ws.cell(row=ri, column=4, value=call_serial2)
    c.number_format = 'HH:MM:SS'

    ws.cell(row=ri, column=5, value=unit)

    c = ws.cell(row=ri, column=6, value=depart_serial)
    c.number_format = 'H:MM:SS'

    ws.cell(row=ri, column=7, value=f'=MINUTE(F{ri}-D{ri})')
    ws.cell(row=ri, column=8, value=None)

# column widths
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

wb.save(OUTPUT)
print(f'Done: {OUTPUT}')
print(f'Rows written: {len(data)} (19.05: {len(df19)}, 20.05: {len(df20)}, after dedup+sort: {len(df_all)})')
