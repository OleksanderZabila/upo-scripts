# -*- coding: utf-8 -*-
import os, sys, json, tkinter as tk, webbrowser
from tkinter import filedialog
import customtkinter as ctk
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── paths ─────────────────────────────────────────────────────────────────────
def _ascii_path(p):
    """Windows short (8.3) path — tksvg/tk fail on non-ASCII characters."""
    try:
        import ctypes
        buf = ctypes.create_unicode_buffer(512)
        if ctypes.windll.kernel32.GetShortPathNameW(p, buf, 512) and buf.value:
            return buf.value
    except Exception:
        pass
    return p

BASE_DIR   = os.path.dirname(os.path.abspath(sys.argv[0]))
BUNDLE_DIR = getattr(sys, '_MEIPASS', BASE_DIR)
EMBLEM_PATH= os.path.join(BUNDLE_DIR, 'upo_emblem.png')
ICO_PATH   = os.path.join(BUNDLE_DIR, 'patrol-polycar.ico')
PNG_PATH   = os.path.join(BUNDLE_DIR, 'patrol-polycar.png')
_APP_DATA  = os.path.join(os.environ.get('APPDATA', BASE_DIR), 'UPOTIMEDRIVER')
os.makedirs(_APP_DATA, exist_ok=True)
CFG_PATH   = os.path.join(_APP_DATA, 'upo_config.json')

UNITS = ['НР 10', 'НР 12', 'НР 13', 'НР 15', 'НР Умань 1', 'НР Умань 2']

APP_VERSION = 'v2.7'
GITHUB_URL  = 'https://github.com/OleksanderZabila/upo-scripts'

# ── palette ───────────────────────────────────────────────────────────────────
BG       = '#0F1923'
CARD     = '#182333'
BORDER   = '#243550'
BTN      = '#1B55A8'
BTN_HOV  = '#144090'
GOLD     = '#D4A82A'
WHITE    = '#FFFFFF'
TEXT     = '#CDD9EE'
MUTED    = '#5E7FA8'
ENTRY_BG = '#0B1420'
SUCCESS  = '#2EA84D'
WARN     = '#D08C00'
ERR      = '#C43030'
LINK     = '#5C9FD6'
ALT_ROW  = 'B0C4DE'
FOOTER   = '#080F18'

S_COL = ['#1A6B30', '#1A4FA0', '#9A3D00', '#8A1212', '#491080']
# pastel duration-bucket fills (Excel cells) — same hues as S_COL, much lighter
DUR_FILL_HEX = ['C8E6CF', 'C9D9F0', 'F5D6B0', 'F0C8C8', 'DCC4EE']

ctk.set_appearance_mode('dark')

# ── config ────────────────────────────────────────────────────────────────────
def load_cfg():
    # migrate from old path
    old = os.path.join(os.environ.get('APPDATA', BASE_DIR), 'УПО Scripts', 'upo_config.json')
    if os.path.exists(old) and not os.path.exists(CFG_PATH):
        try:
            import shutil
            shutil.copy2(old, CFG_PATH)
        except Exception:
            pass
    if os.path.exists(CFG_PATH):
        try:
            with open(CFG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {'car_numbers': {u: '' for u in UNITS}}

def save_cfg(cfg):
    with open(CFG_PATH, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

# ── helpers ───────────────────────────────────────────────────────────────────
def td_serial(td):
    return td.total_seconds() / 86400 if hasattr(td, 'total_seconds') else None

def get_car(unit, car_map):
    s = str(unit).strip()
    if s in car_map:
        return car_map[s]
    for k, v in car_map.items():
        if k.strip().lower() == s.lower():
            return v
    return ''

def load_frames(paths):
    frames = []
    for p in paths:
        if not p or not os.path.exists(p):
            continue
        try:
            xl = pd.ExcelFile(p)
            for sh_idx, sh_name in enumerate(xl.sheet_names[:2]):
                try:
                    df = xl.parse(sh_idx, header=None)
                    df['_date'] = sh_name
                    frames.append(df)
                except Exception:
                    pass
        except Exception:
            pass
    if not frames:
        return None

    df = pd.concat(frames, ignore_index=True)

    def valid(row):
        try:
            return (
                hasattr(row[1], 'total_seconds') and   # call B
                hasattr(row[6], 'total_seconds') and   # arrival G
                pd.notna(row[3]) and str(row[3]).strip() != '' and
                pd.notna(row[5]) and str(row[5]).strip() != ''
            )
        except Exception:
            return False

    df = df[df.apply(valid, axis=1)]
    df = df.drop_duplicates(subset=['_date', 1, 2])

    # sort key: (date from sheet name, time-of-day in seconds)
    def parse_sheet_date(s):
        s = str(s).strip()
        for fmt in ('%d.%m.%Y', '%d.%m.%y', '%d.%m'):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                continue
        return datetime.max  # unknown → push to end

    df['_dk'] = df['_date'].apply(parse_sheet_date)
    df['_tk'] = df[1].apply(lambda v: v.total_seconds() if hasattr(v, 'total_seconds') else 0)
    return df.sort_values(['_dk', '_tk']).drop(columns=['_dk', '_tk']).reset_index(drop=True)

# ── statistics ────────────────────────────────────────────────────────────────
def _bin_durations(durs):
    if not durs:
        return None
    return {
        'до_3':  sum(1 for x in durs if x <= 3),
        '3_6':   sum(1 for x in durs if 3  < x <= 6),
        '6_9':   sum(1 for x in durs if 6  < x <= 9),
        '9_15':  sum(1 for x in durs if 9  < x <= 15),
        '15p':   sum(1 for x in durs if x  > 15),
        'total': len(durs),
    }

def calc_stats_from_formatted(path):
    """Read a previously-generated output xlsx and bin its arrival durations.
    Computes (col 7 'Час відбуття' − col 5 'Час прийому виклику') in minutes,
    so it works even if the formula in col 8 was never evaluated by Excel."""
    try:
        df = pd.read_excel(path, header=0)
    except Exception:
        return None
    if df is None or df.empty or df.shape[1] < 7:
        return None

    def to_sec(v):
        if v is None:
            return None
        if hasattr(v, 'total_seconds'):
            return v.total_seconds()
        if hasattr(v, 'hour') and hasattr(v, 'minute'):  # datetime.time / Timestamp
            return v.hour * 3600 + v.minute * 60 + getattr(v, 'second', 0)
        if isinstance(v, (int, float)):
            if 0 <= v < 1:                # serial fraction-of-day
                return v * 86400
            return v
        try:
            s = str(v).strip()
            parts = s.split(':')
            if len(parts) >= 2:
                h = int(parts[0]); m = int(parts[1])
                sec = int(parts[2]) if len(parts) > 2 else 0
                return h*3600 + m*60 + sec
        except Exception:
            pass
        return None

    durs = []
    e_col = df.columns[4]    # col 5 — Час прийому виклику
    g_col = df.columns[6]    # col 7 — Час відбуття
    for _, row in df.iterrows():
        es = to_sec(row[e_col])
        gs = to_sec(row[g_col])
        if es is None or gs is None:
            continue
        d = (gs - es) / 60
        if d >= 0:
            durs.append(d)
    return _bin_durations(durs)

def calc_stats(df):
    if df is None or df.empty:
        return None
    durs = []
    for _, row in df.iterrows():
        b, g = row[1], row[6]
        if hasattr(b, 'total_seconds') and hasattr(g, 'total_seconds'):
            d = (g.total_seconds() - b.total_seconds()) / 60
            if d >= 0:
                durs.append(d)
    if not durs:
        return None
    return {
        'до_3':  sum(1 for x in durs if x <= 3),
        '3_6':   sum(1 for x in durs if 3  < x <= 6),
        '6_9':   sum(1 for x in durs if 6  < x <= 9),
        '9_15':  sum(1 for x in durs if 9  < x <= 15),
        '15p':   sum(1 for x in durs if x  > 15),
        'total': len(durs),
    }

# ── conversion ────────────────────────────────────────────────────────────────
def convert(df, car_map, colorize=False, src_path=''):
    wb = Workbook()
    ws = wb.active
    try:
        ws.title = datetime.today().strftime('%d.%m.%y')
    except Exception:
        pass

    HDR = ['Дата', 'ТП', 'Назва', 'Адреса', 'Час прийому виклику',
           'Наряд', 'Час прибуття', 'Тривалість (хв)', 'Номер авто НР', 'Опис']

    thin = Side(style='thin', color='8899AA')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf   = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    hfil = PatternFill('solid', fgColor='1B3865')
    hal  = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ci, h in enumerate(HDR, 1):
        c = ws.cell(1, ci, h)
        c.font = hf; c.fill = hfil; c.alignment = hal; c.border = brd
    ws.row_dimensions[1].height = 30

    rf   = Font(name='Arial', size=10)
    cen  = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left',   vertical='center')
    afil = PatternFill('solid', fgColor=ALT_ROW)

    dur_fills = [PatternFill('solid', fgColor=h) for h in DUR_FILL_HEX]

    def dur_bucket_fill(mins):
        if mins is None or mins < 0:
            return None
        if mins <= 3:  return dur_fills[0]
        if mins <= 6:  return dur_fills[1]
        if mins <= 9:  return dur_fills[2]
        if mins <= 15: return dur_fills[3]
        return dur_fills[4]

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        try:
            cs  = td_serial(row[1])      # call reception (input B)
            ds  = td_serial(row[6])      # arrival       (input G)
            car = get_car(row[5], car_map)
            date_val = str(row.get('_date', '')).strip()
            dur_mins = (row[6].total_seconds() - row[1].total_seconds()) / 60
        except Exception:
            continue

        def wr(col, val, fmt=None, al=cen):
            c = ws.cell(ri, col, val)
            c.font = rf; c.border = brd; c.alignment = al
            if fmt: c.number_format = fmt
            if colorize and ri % 2 == 0: c.fill = afil
            return c

        # опис виклику з вхідної колонки J (індекс 9), якщо є
        try:
            opys = row[9]
            opys = '' if pd.isna(opys) else str(opys).strip()
        except Exception:
            opys = ''

        wr(1, date_val)                          # Дата
        wr(2, cs,             'H:MM:SS')         # ТП                  = input B
        wr(3, row[2],         al=left)           # Назва               = input C
        wr(4, row[3],         al=left)           # Адреса              = input D
        wr(5, cs,             'HH:MM:SS')        # Час прийому виклику = input B
        wr(6, row[5])                            # Наряд               = input F
        wr(7, ds,             'H:MM:SS')         # Час прибуття        = input G
        dur_cell = wr(8, f'=MINUTE(G{ri}-E{ri})')  # Тривалість = G − E
        wr(9, car)                               # Номер авто НР
        wr(10, opys,          al=left)           # Опис                = input J

        # always color duration cell by bucket (independent of row alt-colorize)
        fill = dur_bucket_fill(dur_mins)
        if fill is not None:
            dur_cell.fill = fill

    for i, w in enumerate([12, 10, 30, 34, 12, 12, 10, 12, 18, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    base    = src_path or BASE_DIR
    out_dir = os.path.dirname(base) if os.path.isfile(base) else base
    date_str = datetime.today().strftime('%d.%m.%Y')
    out = os.path.join(out_dir, f'УПО_{date_str}.xlsx')
    n = 1
    while os.path.exists(out):
        out = os.path.join(out_dir, f'УПО_{date_str}_{n}.xlsx'); n += 1
    wb.save(out)
    return out, len(df)

# ── stats panel ───────────────────────────────────────────────────────────────
STAT_LABELS = ['До 3 хв', '3.01–6 хв', '6.01–9 хв', '9.01–15 хв', 'Більше 15 хв']
STAT_KEYS   = ['до_3', '3_6', '6_9', '9_15', '15p']

class StatsPanel(ctk.CTkFrame):
    def __init__(self, parent, **kw):
        super().__init__(parent, fg_color='transparent', **kw)
        self._nums  = []
        self._total = None
        self._build()

    def _build(self):
        for ci, (label, color) in enumerate(zip(STAT_LABELS, S_COL)):
            card = ctk.CTkFrame(self, fg_color=color, corner_radius=8,
                                width=125, height=80)
            card.grid(row=0, column=ci, padx=5, pady=0)
            card.pack_propagate(False)

            num = ctk.CTkLabel(card, text='—',
                               font=('Arial', 28, 'bold'),
                               text_color=WHITE, fg_color='transparent')
            num.pack(expand=True, pady=(10, 2))
            self._nums.append(num)

            ctk.CTkLabel(card, text=label,
                         font=('Arial', 9), text_color='#C0D8FF',
                         fg_color='transparent').pack(pady=(0, 10))

        tot = ctk.CTkFrame(self, fg_color='transparent')
        tot.grid(row=1, column=0, columnspan=5, sticky='e', pady=(10, 0))

        ctk.CTkLabel(tot, text='Всього:',
                     font=('Arial', 12), text_color=MUTED,
                     fg_color='transparent').pack(side='left', padx=(0, 6))

        self._total = ctk.CTkLabel(tot, text='—',
                                   font=('Arial', 20, 'bold'),
                                   text_color=GOLD, fg_color='transparent')
        self._total.pack(side='left')

    def update(self, stats):
        if stats is None:
            for n in self._nums: n.configure(text='—')
            self._total.configure(text='—')
            return
        for n, key in zip(self._nums, STAT_KEYS):
            n.configure(text=str(stats[key]))
        self._total.configure(text=str(stats['total']))

# ── file row widget ───────────────────────────────────────────────────────────
class FileRow(ctk.CTkFrame):
    def __init__(self, parent, label, on_change=None, **kw):
        super().__init__(parent, fg_color='transparent', **kw)
        self.var       = tk.StringVar()
        self.enabled   = tk.BooleanVar(value=True)
        self.on_change = on_change

        ctk.CTkCheckBox(self, text='', variable=self.enabled, width=28,
                        fg_color=BTN, hover_color=BTN_HOV,
                        border_color=BORDER, checkmark_color=WHITE,
                        command=self._changed).pack(side='left')

        ctk.CTkLabel(self, text=label, width=58, anchor='w',
                     font=('Arial', 11), text_color=MUTED,
                     fg_color='transparent').pack(side='left')

        ctk.CTkEntry(self, textvariable=self.var, width=314, height=32,
                     fg_color=ENTRY_BG, border_color=BORDER,
                     text_color=WHITE, font=('Arial', 11),
                     placeholder_text='Оберіть .xlsx файл...'
                     ).pack(side='left', padx=(0, 6))

        ctk.CTkButton(self, text='📂', width=44, height=32,
                      fg_color=BTN, hover_color=BTN_HOV,
                      text_color=WHITE, font=('Arial', 13),
                      corner_radius=6, command=self._pick
                      ).pack(side='left', padx=(0, 4))

        ctk.CTkButton(self, text='✕', width=32, height=32,
                      fg_color='#7A2222', hover_color='#9A2A2A',
                      text_color=WHITE, font=('Arial', 13, 'bold'),
                      corner_radius=6, command=self._clear
                      ).pack(side='left')

        self.var.trace_add('write', lambda *_: self._changed())

    def _pick(self):
        p = filedialog.askopenfilename(
            title='Оберіть Excel файл',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Всі файли', '*.*')])
        if p:
            self.var.set(p)

    def _clear(self):
        self.var.set('')
        self.enabled.set(True)

    def _changed(self):
        if self.on_change:
            self.on_change()

    def get(self):
        p = self.var.get().strip()
        return p if (self.enabled.get() and p and os.path.exists(p)) else None

# ── settings dialog ───────────────────────────────────────────────────────────
class SettingsDialog(ctk.CTkToplevel):
    def __init__(self, parent, cfg, on_save):
        super().__init__(parent)
        self.title('Номери автомобілів')
        self.geometry('440x390')
        self.resizable(False, False)
        self.configure(fg_color=BG)
        self.grab_set()
        self.cfg     = cfg
        self.on_save = on_save
        self.entries = {}
        try: self.iconbitmap(ICO_PATH)
        except Exception: pass

        ctk.CTkLabel(self, text='Номери автомобілів по підрозділах',
                     font=('Arial', 14, 'bold'), text_color=GOLD
                     ).pack(pady=(18, 10))

        fr = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10,
                          border_width=1, border_color=BORDER)
        fr.pack(padx=20, fill='x')

        for u in UNITS:
            row = ctk.CTkFrame(fr, fg_color='transparent')
            row.pack(fill='x', padx=14, pady=6)
            ctk.CTkLabel(row, text=u, width=120, anchor='w',
                         font=('Arial', 12), text_color=TEXT
                         ).pack(side='left')
            e = ctk.CTkEntry(row, width=220, fg_color=ENTRY_BG,
                             border_color=BORDER, text_color=WHITE,
                             font=('Arial', 12), placeholder_text='держномер...')
            val = cfg.get('car_numbers', {}).get(u, '')
            if val:
                e.insert(0, val)
            e.pack(side='left', padx=(8, 0))
            self.entries[u] = e

        ctk.CTkButton(self, text='Зберегти',
                      fg_color=BTN, hover_color=BTN_HOV,
                      text_color=WHITE, font=('Arial', 13, 'bold'),
                      corner_radius=8, height=38, command=self._save
                      ).pack(pady=16)

    def _save(self):
        self.cfg['car_numbers'] = {u: self.entries[u].get().strip() for u in UNITS}
        save_cfg(self.cfg)
        self.on_save(self.cfg['car_numbers'])
        self.destroy()

# ── main window ───────────────────────────────────────────────────────────────
W, H = 720, 768

class UPOApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title('UPOTIMEDRIVER')
        self.geometry(f'{W}x{H}')
        self.resizable(False, False)
        self.configure(fg_color=BG)
        try: self.iconbitmap(ICO_PATH)
        except Exception: pass

        self.cfg      = load_cfg()
        self.car_map  = self.cfg.get('car_numbers', {})
        self.colorize = tk.BooleanVar(value=False)

        self._build_bg()
        self._build_ui()

    # ── background canvas ─────────────────────────────────────────────────────
    def _build_bg(self):
        c = tk.Canvas(self, bg=BG, highlightthickness=0)
        c.place(x=0, y=0, width=W, height=H)

        # accent lines
        c.create_rectangle(0, 0, W, 3,     fill=GOLD,   outline='')
        c.create_rectangle(0, H-42, W, H-2, fill='#0E1828', outline='')
        c.create_rectangle(0, H-43, W, H-42, fill=BORDER, outline='')
        c.create_rectangle(0, H-2, W, H,    fill=BTN,    outline='')

    # ── main UI ───────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── header ────────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color='transparent', width=W, height=120)
        hdr.place(x=0, y=6)
        hdr.pack_propagate(False)

        # patrol car — left
        if os.path.exists(PNG_PATH):
            try:
                raw = Image.open(PNG_PATH).convert('RGBA')
                raw.thumbnail((220, 108), Image.LANCZOS)
                self._car = ctk.CTkImage(light_image=raw, dark_image=raw,
                                         size=(raw.width, raw.height))
                ctk.CTkLabel(hdr, image=self._car, text='',
                             fg_color='transparent').place(x=12, y=8)
            except Exception:
                pass

        # title + subtitle — center
        ctk.CTkLabel(hdr, text='UPOTIMEDRIVER',
                     font=('Arial', 28, 'bold'), text_color=GOLD,
                     fg_color='transparent').place(relx=0.5, y=22, anchor='n')
        ctk.CTkLabel(hdr, text='Обробка файлів виїздів',
                     font=('Arial', 11), text_color=MUTED,
                     fg_color='transparent').place(relx=0.5, y=65, anchor='n')

        # UPO emblem — right side of header (PNG via CTkImage)
        if os.path.exists(EMBLEM_PATH):
            try:
                emb = Image.open(EMBLEM_PATH).convert('RGBA')
                emb.thumbnail((100, 100), Image.LANCZOS)
                self._hdr_emb = ctk.CTkImage(light_image=emb, dark_image=emb,
                                             size=(emb.width, emb.height))
                ctk.CTkLabel(hdr, image=self._hdr_emb, text='',
                             fg_color='transparent'
                             ).place(x=W - 18 - emb.width, y=10)
            except Exception:
                pass

        sep = tk.Canvas(self, bg=BORDER, highlightthickness=0, height=1)
        sep.place(x=17, y=128, width=W - 34)

        # ── files card ────────────────────────────────────────────────────────
        fc = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10,
                          border_width=1, border_color=BORDER,
                          width=686, height=160)
        fc.place(relx=0.5, y=142, anchor='n')
        fc.pack_propagate(False)

        ctk.CTkLabel(fc, text='Вхідні файли  (☑ — обробляти)',
                     font=('Arial', 10), text_color=MUTED,
                     fg_color='transparent').place(x=14, y=7)

        self.file1 = FileRow(fc, 'Файл 1:', on_change=self._refresh_stats)
        self.file1.place(x=10, y=30)
        self.file2 = FileRow(fc, 'Файл 2:', on_change=self._refresh_stats)
        self.file2.place(x=10, y=70)

        ctk.CTkLabel(fc, text='Готовий файл — лише для перегляду статистики',
                     font=('Arial', 9), text_color=GOLD,
                     fg_color='transparent').place(x=14, y=112)
        self.file3 = FileRow(fc, 'Готовий:', on_change=self._refresh_stats)
        self.file3.place(x=10, y=124)

        # ── stats card ────────────────────────────────────────────────────────
        sc = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10,
                          border_width=1, border_color=BORDER,
                          width=686, height=144)
        sc.place(relx=0.5, y=316, anchor='n')
        sc.pack_propagate(False)

        ctk.CTkLabel(sc, text='Статистика тривалості прибуття',
                     font=('Arial', 10), text_color=MUTED,
                     fg_color='transparent').place(x=14, y=8)

        self.stats_panel = StatsPanel(sc)
        self.stats_panel.place(x=8, y=28)

        # ── settings card ─────────────────────────────────────────────────────
        nc = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10,
                          border_width=1, border_color=BORDER,
                          width=686, height=62)
        nc.place(relx=0.5, y=474, anchor='n')
        nc.pack_propagate(False)

        ctk.CTkLabel(nc, text='Номери автомобілів',
                     font=('Arial', 11), text_color=TEXT,
                     fg_color='transparent').place(x=14, y=8)
        self.lbl_cars = ctk.CTkLabel(nc, text=self._summary(),
                                     font=('Arial', 10), text_color=MUTED,
                                     fg_color='transparent')
        self.lbl_cars.place(x=14, y=34)
        ctk.CTkButton(nc, text='⚙  Змінити', width=100, height=30,
                      fg_color=BTN, hover_color=BTN_HOV,
                      text_color=WHITE, font=('Arial', 11),
                      corner_radius=6, command=self._settings
                      ).place(x=574, y=16)

        # ── colorize toggle ───────────────────────────────────────────────────
        ctk.CTkCheckBox(self, text='Розукрашення рядків',
                        variable=self.colorize,
                        font=('Arial', 11), text_color=MUTED,
                        fg_color=BTN, hover_color=BTN_HOV,
                        border_color=BORDER, checkmark_color=WHITE
                        ).place(relx=0.5, y=558, anchor='center')

        # ── process button ────────────────────────────────────────────────────
        ctk.CTkButton(self, text='▶   Обробити',
                      font=('Arial', 15, 'bold'),
                      fg_color=BTN, hover_color=BTN_HOV,
                      text_color=WHITE, corner_radius=10,
                      height=46, width=220,
                      command=self._process
                      ).place(relx=0.5, y=600, anchor='center')

        # ── open formatted file (review) ──────────────────────────────────────
        ctk.CTkButton(self, text='📋   Переглянути готовий файл',
                      font=('Arial', 11),
                      fg_color='#2A4D78', hover_color='#1F3D60',
                      text_color=TEXT, corner_radius=8,
                      height=32, width=220,
                      command=self._open_output
                      ).place(relx=0.5, y=646, anchor='center')

        # ── status label ──────────────────────────────────────────────────────
        self.lbl_status = ctk.CTkLabel(self, text='',
                                       font=('Arial', 11), text_color=MUTED,
                                       fg_color='transparent', wraplength=660)
        self.lbl_status.place(relx=0.5, y=694, anchor='center')

        # ── footer ────────────────────────────────────────────────────────────
        ctk.CTkLabel(self, text=APP_VERSION,
                     font=('Arial', 10), text_color=MUTED,
                     fg_color='transparent').place(x=16, y=H - 30)

        lnk = ctk.CTkLabel(self, text='github.com/OleksanderZabila/upo-scripts',
                            font=('Arial', 10), text_color=LINK,
                            fg_color='transparent', cursor='hand2')
        lnk.place(relx=0.5, y=H - 30, anchor='n')
        lnk.bind('<Button-1>', lambda _: webbrowser.open(GITHUB_URL))

        ctk.CTkLabel(self, text='UPOTIMEDRIVER',
                     font=('Arial', 9), text_color='#2A3E58',
                     fg_color='transparent').place(x=W - 106, y=H - 30)

    # ── logic ─────────────────────────────────────────────────────────────────
    def _refresh_stats(self):
        # Готовий xlsx (file3) має пріоритет — показує власну статистику
        f3 = self.file3.get()
        if f3:
            stats = calc_stats_from_formatted(f3)
            if stats is not None:
                self.stats_panel.update(stats)
                return

        paths = [f for f in [self.file1.get(), self.file2.get()] if f]
        if not paths:
            self.stats_panel.update(None)
            return
        try:
            df = load_frames(paths)
            self.stats_panel.update(calc_stats(df))
        except Exception:
            self.stats_panel.update(None)

    def _summary(self):
        filled = sum(1 for v in self.car_map.values() if v)
        return f'Заповнено: {filled} / {len(UNITS)}'

    def _settings(self):
        SettingsDialog(self, self.cfg, self._on_saved)

    def _on_saved(self, car_numbers):
        self.car_map = car_numbers
        self.lbl_cars.configure(text=self._summary())
        self._st('Номери авто збережено', SUCCESS)

    def _open_output(self):
        p = filedialog.askopenfilename(
            title='Оберіть готовий xlsx для перегляду',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Всі файли', '*.*')])
        if not p:
            return
        try:
            os.startfile(p)
            self._st(f'Відкрито: {os.path.basename(p)}', SUCCESS)
        except Exception as e:
            self._st(f'Не вдалося відкрити: {e}', ERR)

    def _process(self):
        paths = [f for f in [self.file1.get(), self.file2.get()] if f]
        if not paths:
            self._st('Оберіть та позначте хоча б один файл', WARN)
            return
        self._st('Обробка...', MUTED); self.update()
        try:
            df = load_frames(paths)
            if df is None or df.empty:
                self._st('Файли порожні або не вдалося прочитати', ERR)
                return
            out, n = convert(df, self.car_map, self.colorize.get(), paths[0])
            self._st(f'Готово — {n} рядків → {os.path.basename(out)}', SUCCESS)
        except Exception as e:
            self._st(f'Помилка: {e}', ERR)

    def _st(self, msg, col):
        self.lbl_status.configure(text=msg, text_color=col)


if __name__ == '__main__':
    app = UPOApp()
    app.mainloop()
